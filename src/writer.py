"""
Geração das planilhas de saída:
  1. Atualiza a aba COFINS e PIS do template com os valores apurados
  2. Gera aba de consolidação por NF
  3. Gera aba de validação
"""

import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

from .calculator import ResultadoApuracao
from .validator import Inconsistencia


# ─── Constantes visuais ───────────────────────────────────────────────────────
FONT_NAME     = "Arial"
COLOR_FORMULA = "000000"   # Preto — fórmulas/cálculos
COLOR_INPUT   = "0000FF"   # Azul — valores inseridos
COLOR_HEADER  = "1F3864"   # Azul escuro
COLOR_FORMULA_GREEN = "008000"  # Verde — links inter-abas


# ─── Helpers de localização ───────────────────────────────────────────────────

def _find_row_by_label(sheet, label_partial: str, label_col: int = 3) -> int | None:
    """Localiza linha cujo texto na coluna label_col contém label_partial (case-insensitive)."""
    label_lower = label_partial.lower().strip()
    for row_cells in sheet.iter_rows():
        cell = sheet.cell(row=row_cells[0].row, column=label_col)
        if cell.value and label_lower in str(cell.value).lower():
            return cell.row
    return None


def _find_col_cofins_dataonly(path: Path, sheet_name: str,
                              target_month: int, target_year: int,
                              header_row: int = 9, start_col: int = 7) -> int | None:
    """
    Abre o workbook com data_only=True (resolve fórmulas externas já calculadas)
    para localizar a coluna correta pelo datetime do cabeçalho.
    """
    wb_ro = load_workbook(path, data_only=True)
    sheet = wb_ro[sheet_name]
    for col in range(start_col, sheet.max_column + 1):
        val = sheet.cell(row=header_row, column=col).value
        if isinstance(val, datetime.datetime) and val.month == target_month and val.year == target_year:
            return col
    return None


def _find_col_pis(sheet, target_month: int, target_year: int) -> int | None:
    """
    Na aba PIS as datas do cabeçalho são fórmulas externas (não resolvidas pelo
    openpyxl). A aba tem uma coluna por mês visível; a última coluna corresponde
    ao mês mais recente. Usa a última coluna com dados na linha de receita (row 11).
    Se não achar por datetime, retorna a última coluna com dados numéricos.
    """
    # Tenta localizar por datetime (funciona se o arquivo foi aberto/salvo no Excel)
    for row_idx in [7, 9, 5, 6]:
        for col in range(7, sheet.max_column + 1):
            val = sheet.cell(row=row_idx, column=col).value
            if isinstance(val, datetime.datetime) and val.month == target_month and val.year == target_year:
                return col

    # Fallback: última coluna que tem dado numérico na linha de receita (row 11)
    receita_row = _find_row_by_label(sheet, "regime de caixa") or 11
    last_col = None
    for col in range(7, sheet.max_column + 1):
        val = sheet.cell(row=receita_row, column=col).value
        if val is not None and val != 0:
            last_col = col
    if last_col:
        return last_col

    # Último recurso: coluna máxima
    return sheet.max_column


def _write_value(sheet, row: int, col: int, value: float):
    """Escreve valor numérico com formatação de input (azul)."""
    cell = sheet.cell(row=row, column=col)
    cell.value = round(value, 2)
    cell.font = Font(name=FONT_NAME, color=COLOR_INPUT, size=10)
    cell.number_format = '#,##0.00'


# ─── Atualização das abas do template ────────────────────────────────────────

def _atualizar_aba_cofins(sheet, resultado: ResultadoApuracao,
                           target_month: int, target_year: int, col: int):
    c = resultado.cofins

    if col is None:
        raise ValueError(
            f"Coluna {target_month:02d}/{target_year} não encontrada na aba COFINS."
        )

    # Mapeamento: label parcial → valor a inserir
    # A lógica busca o label em col C e insere o valor na coluna do mês
    # base_calculo = recebido + retentions - juros (reconstituição da receita bruta)
    mappings = {
        "fluxus":                 c.base_calculo,   # Base líquida na linha de receita
        "retida fonte p/ outras": c.retencao_fonte, # COFINS retida fonte
    }

    inserted = {}
    for label, value in mappings.items():
        row = _find_row_by_label(sheet, label)
        if row:
            _write_value(sheet, row, col, value)
            inserted[label] = (row, value)

    return inserted, col


def _atualizar_aba_pis(sheet, resultado: ResultadoApuracao,
                        target_month: int, target_year: int, col: int):
    p = resultado.pis

    if col is None:
        raise ValueError(f"Coluna {target_month:02d}/{target_year} não encontrada na aba PIS.")

    mappings = {
        "fluxus":       p.base_calculo,    # Base líquida na linha de receita
        "retida fonte": p.retencao_fonte,
    }

    inserted = {}
    for label, value in mappings.items():
        row = _find_row_by_label(sheet, label)
        if row:
            _write_value(sheet, row, col, value)
            inserted[label] = (row, value)

    return inserted, col


# ─── Aba de consolidação por NF ───────────────────────────────────────────────

def _criar_aba_consolidacao(wb, dados: dict[str, pd.DataFrame], competencia: str):
    if "Consolidação NF" in wb.sheetnames:
        del wb["Consolidação NF"]
    ws = wb.create_sheet("Consolidação NF")

    base = dados["recebidas"].copy()
    for key, col in [
        ("cofins_ret", "cofins_retido"),
        ("pis_ret",    "pis_retido"),
        ("csll_ret",   "csll_retido"),
        ("irrf",       "irrf"),
        ("juros",      "juros"),
    ]:
        df = dados[key]
        if not df.empty:
            base = base.merge(df, on="nf", how="outer")

    num_cols = ["recebido", "cofins_retido", "pis_retido", "csll_retido", "irrf", "juros"]
    for col in num_cols:
        if col in base.columns:
            base[col] = pd.to_numeric(base[col], errors="coerce").fillna(0)
        else:
            base[col] = 0.0

    base = base.sort_values("nf").reset_index(drop=True)

    headers = [
        "NF", "Cliente",
        "Valor Recebido (R$)", "COFINS Retido (R$)", "PIS Retido (R$)",
        "CSLL Retida (R$)", "IRRF (R$)", "Juros/Multas (R$)",
        "Total Deduções (R$)", "Base Líquida (R$)",
    ]

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    t = ws["A1"]
    t.value = f"CONSOLIDAÇÃO POR NOTA FISCAL — COMPETÊNCIA {competencia}"
    t.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=12)
    t.fill = PatternFill("solid", start_color=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Cabeçalhos
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", start_color="2E75B6")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Dados
    num_format = '#,##0.00'
    for ri, row_data in base.iterrows():
        excel_row = ri + 3
        fill_color = "F2F2F2" if ri % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", start_color=fill_color)

        ws.cell(row=excel_row, column=1, value=str(row_data.get("nf", ""))).font = Font(name=FONT_NAME, size=9)
        ws.cell(row=excel_row, column=2, value=str(row_data.get("cliente", ""))).font = Font(name=FONT_NAME, size=9)

        for ci, key in enumerate(["recebido", "cofins_retido", "pis_retido", "csll_retido", "irrf", "juros"], 3):
            c = ws.cell(row=excel_row, column=ci, value=round(float(row_data.get(key, 0)), 2))
            c.number_format = num_format
            c.font = Font(name=FONT_NAME, size=9)
            c.fill = fill

        # Total deduções (fórmula)
        c_tot = ws.cell(row=excel_row, column=9,
                         value=f"=SUM(D{excel_row}:G{excel_row})")
        c_tot.number_format = num_format
        c_tot.font = Font(name=FONT_NAME, size=9, color=COLOR_FORMULA)
        c_tot.fill = fill

        # Base líquida = Recebido + COFINS + PIS + CSLL + IRRF - Juros
        c_base = ws.cell(row=excel_row, column=10,
                          value=f"=C{excel_row}+D{excel_row}+E{excel_row}+F{excel_row}+G{excel_row}-H{excel_row}")
        c_base.number_format = num_format
        c_base.font = Font(name=FONT_NAME, size=9, color=COLOR_FORMULA)
        c_base.fill = fill

    # Linha de totais
    total_row = len(base) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name=FONT_NAME, bold=True, size=10)
    ws.cell(row=total_row, column=2, value="").font = Font(name=FONT_NAME, size=9)
    total_fill = PatternFill("solid", start_color="D9E1F2")
    for ci in range(3, 11):
        ltr = get_column_letter(ci)
        c = ws.cell(row=total_row, column=ci, value=f"=SUM({ltr}3:{ltr}{total_row-1})")
        c.number_format = num_format
        c.font = Font(name=FONT_NAME, bold=True, size=10, color=COLOR_FORMULA)
        c.fill = total_fill

    # Larguras
    for i, w in enumerate([14, 42, 20, 18, 14, 14, 14, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"


# ─── Aba de validação / resumo ────────────────────────────────────────────────

def _criar_aba_validacao(wb, alertas: list[Inconsistencia], resultado: ResultadoApuracao):
    if "Validação" in wb.sheetnames:
        del wb["Validação"]
    ws = wb.create_sheet("Validação")

    c = resultado.cofins
    p = resultado.pis

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"VALIDAÇÃO E RESUMO — COMPETÊNCIA {resultado.competencia}"
    t.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=12)
    t.fill = PatternFill("solid", start_color=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    resumo = [
        ("RESUMO DA APURAÇÃO", "", True),
        ("Competência", resultado.competencia, False),
        ("", "", False),
        ("(+) Total Recebido de Clientes", resultado.total_recebido, False),
        (f"(+) COFINS Retido na Fonte", c.retencao_fonte, False),
        (f"(+) PIS Retido na Fonte", p.retencao_fonte, False),
        ("(+) CSLL Retida (reconstituição)", resultado.csll_retida, False),
        ("(+) IRRF Retido (reconstituição)", resultado.irrf_retido, False),
        ("(-) Juros/Multas Recebidos", resultado.total_juros, False),
        ("BASE LÍQUIDA (PIS e COFINS)", c.base_calculo, True),
        ("", "", False),
        (f"COFINS — Alíquota {c.aliquota*100:.2f}%", "", True),
        ("  Valor Apurado", c.valor_apurado, False),
        ("  Retido na Fonte", c.retencao_fonte, False),
        ("  COFINS A RECOLHER", c.valor_a_pagar, True),
        ("", "", False),
        (f"PIS — Alíquota {p.aliquota*100:.4f}%", "", True),
        ("  Valor Apurado", p.valor_apurado, False),
        ("  Retido na Fonte", p.retencao_fonte, False),
        ("  PIS A RECOLHER", p.valor_a_pagar, True),
        ("", "", False),
        ("CSLL Retida (informativo)", resultado.csll_retida, False),
        ("IRRF Retido (informativo)", resultado.irrf_retido, False),
    ]

    pagar_labels = {"  COFINS A RECOLHER", "  PIS A RECOLHER"}

    for i, (label, value, bold) in enumerate(resumo, 2):
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2)

        lc.font = Font(name=FONT_NAME, bold=bold, size=10)

        if isinstance(value, float):
            vc.value = round(value, 2)
            vc.number_format = 'R$ #,##0.00'
            if label in pagar_labels:
                vc.font = Font(name=FONT_NAME, bold=True, size=11)
                vc.fill = PatternFill("solid", start_color="E2EFDA")
                lc.fill = PatternFill("solid", start_color="E2EFDA")
                lc.font = Font(name=FONT_NAME, bold=True, size=11)
            else:
                vc.font = Font(name=FONT_NAME, size=10)
        else:
            vc.value = value
            vc.font = Font(name=FONT_NAME, bold=bold, size=10)

    # Alertas
    alert_start = len(resumo) + 4
    ws.cell(row=alert_start, column=1, value="ALERTAS DE VALIDAÇÃO").font = Font(
        name=FONT_NAME, bold=True, size=11)

    if not alertas:
        ws.cell(row=alert_start + 1, column=1,
                value="OK — Nenhuma inconsistencia encontrada.").font = Font(
            name=FONT_NAME, color="008000", size=10)
    else:
        for ai, alerta in enumerate(alertas):
            r = alert_start + 1 + ai * 4
            ws.cell(row=r,   column=1, value=f"[!] {alerta.tipo}").font = Font(
                name=FONT_NAME, bold=True, color="C00000", size=10)
            ws.cell(row=r+1, column=1, value=alerta.descricao).font = Font(name=FONT_NAME, size=9)
            ws.cell(row=r+2, column=1,
                    value=f"Qtd: {alerta.quantidade}   Valor: R$ {alerta.valor_total:,.2f}").font = Font(
                name=FONT_NAME, size=9)
            nfs_str = ", ".join(alerta.nfs[:15])
            if len(alerta.nfs) > 15:
                nfs_str += f" ... (+{len(alerta.nfs)-15})"
            ws.cell(row=r+3, column=1, value=f"NFs: {nfs_str}").font = Font(
                name=FONT_NAME, size=8, color="666666")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22


# ─── Ponto de entrada principal ───────────────────────────────────────────────

def atualizar_template(
    template_path: Path,
    output_path: Path,
    resultado: ResultadoApuracao,
    dados: dict[str, pd.DataFrame],
    target_month: int,
    target_year: int,
    alertas: list[Inconsistencia],
) -> tuple[Path, dict]:
    """
    Carrega o template, preenche a coluna do mês de apuração e adiciona
    abas de consolidação e validação. Retorna (caminho_saída, log_inserções).
    """
    # Passo 1: encontrar colunas via data_only (resolve fórmulas externas já calculadas)
    cofins_col = _find_col_cofins_dataonly(template_path, "COFINS ", target_month, target_year)
    pis_col    = _find_col_pis(load_workbook(template_path, data_only=True)["PIS"],
                                target_month, target_year)

    # Passo 2: abrir normalmente para preservar fórmulas ao salvar
    wb = load_workbook(template_path)

    log = {}

    if "COFINS " in wb.sheetnames:
        ins, col = _atualizar_aba_cofins(wb["COFINS "], resultado, target_month, target_year, cofins_col)
        log["cofins_col"] = col
        log["cofins_insertions"] = ins
    else:
        raise ValueError("Aba 'COFINS ' não encontrada no template.")

    if "PIS" in wb.sheetnames:
        ins, col = _atualizar_aba_pis(wb["PIS"], resultado, target_month, target_year, pis_col)
        log["pis_col"] = col
        log["pis_insertions"] = ins
    else:
        raise ValueError("Aba 'PIS' não encontrada no template.")

    _criar_aba_consolidacao(wb, dados, resultado.competencia)
    _criar_aba_validacao(wb, alertas, resultado)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path, log

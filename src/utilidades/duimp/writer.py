"""
Geração do Excel da DUIMP:
  - aba "Itens": uma linha por mercadoria (dados do produto + peso bruto
    rateado + número da adição inferido).
  - aba "Detalhes": tributos por adição e encargos do processo como estão
    na DI (sem rateio), e a tributação (regime/fundamento) declarada de
    cada item.
  - aba "Resumo": dados gerais do processo.
  - aba "Avisos": quando algo relevante precisa de atenção.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
RATEIO_FILL = PatternFill("solid", fgColor="FFF3CD")
INFERENCIA_FILL = PatternFill("solid", fgColor="DCEEFB")

COLUNAS_ITENS = [
    ("item", "Item", 8, None, None),
    ("numero_adicao", "Número da Adição", 16, None, INFERENCIA_FILL),
    ("part_number", "Código do Produto", 18, None, None),
    ("descricao_complementar", "Descrição do Produto", 55, None, None),
    ("ncm", "NCM", 16, None, None),
    ("quantidade", "Quantidade", 12, "#,##0.00", None),
    ("numero_lote", "Número do Lote", 16, None, None),
    ("data_fabricacao", "Data de Fabricação", 14, None, None),
    ("fabricante_legal", "Fabricante", 30, None, None),
    ("pais_origem", "País de Origem", 20, None, None),
    ("fornecedor", "Fornecedor (Exportador)", 32, None, None),
    ("pais_aquisicao", "País de Aquisição", 20, None, None),
    ("moeda_negociada", "Moeda Negociada", 16, None, None),
    ("valor_unitario", "Valor Unitário", 16, "#,##0.0000", None),
    ("valor_total_venda", "Valor Total (moeda)", 16, "#,##0.00", None),
    ("peso_liquido_kg", "Peso Líquido (kg)", 16, "#,##0.00000", None),
    ("peso_bruto_rateado_kg", "Peso Bruto Rateado (kg)", 18, "#,##0.00000", RATEIO_FILL),
    ("cclasstrib", "Class. Tributária (cClassTrib)", 45, None, None),
    ("finalidade_importacao", "Finalidade da Importação", 22, None, None),
    ("cnpj_destino_final", "CNPJ Destino Final", 18, None, None),
    ("dispositivo_recondicionado", "Dispositivo Recondicionado", 16, None, None),
    ("numero_snvs", "Registro SNVS", 16, None, None),
]
COLUNAS_ITENS_TOTAL = ["quantidade", "peso_liquido_kg", "peso_bruto_rateado_kg", "valor_total_venda"]

COLUNAS_TRIBUTACAO_ITEM = [
    ("item", "Item", 8),
    ("part_number", "Código do Produto", 18),
    ("numero_adicao", "Número da Adição", 14),
    ("pis_regime", "PIS - Regime", 18),
    ("pis_fundamento", "PIS - Fundamento", 45),
    ("cofins_regime", "COFINS - Regime", 18),
    ("cofins_fundamento", "COFINS - Fundamento", 45),
    ("ii_regime", "II - Regime", 22),
    ("ii_fundamento", "II - Fundamento", 40),
    ("ipi_regime", "IPI - Regime", 22),
    ("ipi_fundamento", "IPI - Fundamento", 40),
]


def _cabecalho_tabela(ws, row, colunas_titulos, larguras=None):
    for col_idx, titulo in enumerate(colunas_titulos, start=1):
        cell = ws.cell(row=row, column=col_idx, value=titulo)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if larguras:
            ws.column_dimensions[get_column_letter(col_idx)].width = larguras[col_idx - 1]


def _gerar_aba_itens(ws, itens: list[dict]) -> None:
    for col_idx, (_, titulo, largura, _, fill) in enumerate(COLUNAS_ITENS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.freeze_panes = "A2"

    row_idx = 2
    for item in itens:
        for col_idx, (chave, _, _, formato, _) in enumerate(COLUNAS_ITENS, start=1):
            valor = item.get(chave)
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            if formato:
                cell.number_format = formato
        row_idx += 1

    total_row = row_idx
    idx_por_chave = {chave: i + 1 for i, (chave, *_r) in enumerate(COLUNAS_ITENS)}
    ws.cell(row=total_row, column=1, value=f"TOTAL — {len(itens)} item(ns)").font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    for chave in COLUNAS_ITENS_TOTAL:
        col_idx = idx_por_chave[chave]
        total = sum(item.get(chave) or 0 for item in itens)
        tot_cell = ws.cell(row=total_row, column=col_idx, value=round(total, 5))
        tot_cell.font = Font(bold=True)
        formato = next(f for c, _, _, f, _ in COLUNAS_ITENS if c == chave)
        if formato:
            tot_cell.number_format = formato


def _gerar_aba_detalhes(ws, cabecalho: dict, itens: list[dict]) -> None:
    ws.cell(row=1, column=1, value="Tributos por Adição (como declarado na DI)").font = Font(bold=True, size=12)

    titulos_adicao = ["Adição", "NCM", "II (R$)", "IPI (R$)", "Base de Cálculo PIS/COFINS (R$)", "PIS (R$)", "COFINS (R$)"]
    larguras_adicao = [12, 16, 14, 14, 26, 14, 14]
    row = 3
    _cabecalho_tabela(ws, row, titulos_adicao, larguras_adicao)
    row += 1

    adicoes = cabecalho.get("adicoes") or []
    for a in adicoes:
        valores = [a["numero"], a["ncm"], a.get("ii_valor"), a.get("ipi_valor"),
                   a.get("base_pis_cofins"), a.get("pis_valor"), a.get("cofins_valor")]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=col_idx, value=valor)
            if col_idx >= 3:
                cell.number_format = "#,##0.00"
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col_idx, chave in [(3, "ii_valor"), (4, "ipi_valor"), (5, "base_pis_cofins"),
                            (6, "pis_valor"), (7, "cofins_valor")]:
        total = sum(a.get(chave) or 0 for a in adicoes)
        c = ws.cell(row=total_row, column=col_idx, value=round(total, 2))
        c.font = Font(bold=True)
        c.number_format = "#,##0.00"

    row = total_row + 3
    ws.cell(row=row, column=1, value="Encargos do Processo (como declarado na DI)").font = Font(bold=True, size=12)
    row += 1
    encargos = [
        ("Taxa Siscomex (R$)", cabecalho.get("taxa_siscomex")),
        ("Despesas Aduaneiras (R$)", cabecalho.get("despesas_aduaneiras")),
        ("Frete Internacional (USD)", cabecalho.get("frete_usd")),
        ("Frete Internacional (R$)", cabecalho.get("frete_brl")),
        ("Seguro (USD)", cabecalho.get("seguro_usd")),
        ("Seguro (R$)", cabecalho.get("seguro_brl")),
        ("ICMS Total (R$)", cabecalho.get("icms_total")),
    ]
    for campo, valor in encargos:
        ws.cell(row=row, column=1, value=campo).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=valor)
        if isinstance(valor, (int, float)):
            c.number_format = "#,##0.00"
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Tributação por Item (regime e fundamento declarados na DI)").font = Font(bold=True, size=12)
    row += 2
    titulos_item = [t for _, t, _ in COLUNAS_TRIBUTACAO_ITEM]
    larguras_item = [l for _, _, l in COLUNAS_TRIBUTACAO_ITEM]
    _cabecalho_tabela(ws, row, titulos_item, larguras_item)
    row += 1
    for item in itens:
        for col_idx, (chave, _, _) in enumerate(COLUNAS_TRIBUTACAO_ITEM, start=1):
            ws.cell(row=row, column=col_idx, value=item.get(chave))
        row += 1


def _gerar_aba_resumo(ws, cabecalho: dict, qtd_itens: int) -> None:
    linhas = [
        ("Número da DUIMP", cabecalho.get("duimp_numero")),
        ("Versão", cabecalho.get("versao")),
        ("Importador — CNPJ", cabecalho.get("importador_cnpj")),
        ("Importador — Nome", cabecalho.get("importador_nome")),
        ("Processo Virtua Comex", cabecalho.get("processo_virtua")),
        ("Processo Biomedical", cabecalho.get("processo_biomedical")),
        ("Exportador / Fabricante", cabecalho.get("exportador_fabricante")),
        ("Fatura Comercial", cabecalho.get("fatura_comercial")),
        ("Data de Embarque", cabecalho.get("data_embarque")),
        ("Data de Chegada", cabecalho.get("data_chegada")),
        ("Câmbio (Taxa Dólar)", cabecalho.get("taxa_dolar")),
        ("FOB (USD)", cabecalho.get("fob_usd")),
        ("FOB (R$)", cabecalho.get("fob_brl")),
        ("Valor Aduaneiro (USD)", cabecalho.get("valor_aduaneiro_usd")),
        ("Valor Aduaneiro (R$)", cabecalho.get("valor_aduaneiro_brl")),
        ("Peso Bruto Total (kg)", cabecalho.get("peso_bruto_total_kg")),
        ("Peso Líquido Total (kg)", cabecalho.get("peso_liquido_total_kg")),
        ("Quantidade de Itens", qtd_itens),
        ("Quantidade de Adições", len(cabecalho.get("adicoes") or [])),
    ]

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    for row_idx, (campo, valor) in enumerate(linhas, start=1):
        c1 = ws.cell(row=row_idx, column=1, value=campo)
        c1.font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=valor)

    nota_row = len(linhas) + 2
    ws.cell(row=nota_row, column=1, value="Metodologia").font = Font(bold=True, size=12)
    nota = (
        "Tributos (II, IPI, PIS, COFINS) e encargos (frete, seguro, taxa Siscomex, despesas "
        "aduaneiras) NÃO são rateados por item — estão na aba \"Detalhes\" exatamente como "
        "declarados na DI (por adição/processo, não por mercadoria). O único valor rateado por "
        "item é o PESO BRUTO (destacado em amarelo na aba Itens), proporcionalmente à participação "
        "de cada item no peso líquido total — é uma estimativa, já que a DUIMP só informa o peso "
        "bruto total do processo. A coluna \"Número da Adição\" (destacada em azul na aba Itens) é "
        "uma inferência por agrupamento de fabricante, não um dado declarado por item — confira a "
        "aba Avisos."
    )
    ws.cell(row=nota_row + 1, column=1, value=nota).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=nota_row + 1, start_column=1, end_row=nota_row + 1, end_column=2)
    ws.row_dimensions[nota_row + 1].height = 75


def gerar_excel(cabecalho: dict, itens: list[dict], avisos: list[str]) -> bytes:
    wb = Workbook()
    ws_itens = wb.active
    ws_itens.title = "Itens"
    _gerar_aba_itens(ws_itens, itens)

    ws_detalhes = wb.create_sheet("Detalhes")
    _gerar_aba_detalhes(ws_detalhes, cabecalho, itens)

    ws_resumo = wb.create_sheet("Resumo")
    _gerar_aba_resumo(ws_resumo, cabecalho, len(itens))

    if avisos:
        ws_avisos = wb.create_sheet("Avisos")
        ws_avisos.cell(row=1, column=1, value="Avisos do processamento").font = Font(bold=True, size=12)
        for i, aviso in enumerate(avisos, start=3):
            ws_avisos.cell(row=i, column=1, value=aviso)
        ws_avisos.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

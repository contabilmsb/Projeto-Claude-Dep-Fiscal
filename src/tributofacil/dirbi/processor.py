"""
Módulo Dirbi: insere na planilha mensal do Dirbi as colunas de cálculo de
PIS/COFINS (PIS %, COFINS %, PIS 2,1, COFINS 9,65, PIS, COFINS), replicando
exatamente as fórmulas do arquivo de referência (mês em que a planilha já
sai pronta da empresa).

As colunas são localizadas pelo nome do cabeçalho, não pela posição — os
arquivos de meses diferentes podem ter colunas extras/faltantes antes da
"Base Cálculo PIS/COFINS", então a posição das colunas de imposto pode
variar de um arquivo para outro.
"""

import io
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.utils import get_column_letter

HEADER_BASE_CALCULO = "Base Cálculo PIS/COFINS"
HEADER_PIS_TOTAL = "PIS Total(R$)"
HEADER_COFINS_TOTAL = "COFINS Total(R$)"

FILL_PERCENTUAL = PatternFill(patternType="solid", fgColor="FFFFFF00")
FILL_ALIQUOTA = PatternFill(patternType="solid", fgColor=Color(theme=4, tint=0.7999816888943144))
FILL_RESULTADO = PatternFill(patternType="solid", fgColor="FF00B050")
FORMATO_PERCENTUAL = "0.000%"


def _encontrar_coluna(ws, header: str) -> int:
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == header:
            return c
    raise ValueError(
        f'coluna "{header}" não encontrada na planilha. '
        f"Confira se a primeira linha tem os cabeçalhos esperados."
    )


def processar(conteudo: bytes) -> tuple[bytes, dict]:
    wb = load_workbook(io.BytesIO(conteudo), data_only=False)
    ws = wb.worksheets[0]

    if ws.max_row < 2:
        raise ValueError("A planilha não tem nenhuma linha de dados.")

    idx_p = _encontrar_coluna(ws, HEADER_BASE_CALCULO)
    idx_s = _encontrar_coluna(ws, HEADER_PIS_TOTAL)
    idx_t = _encontrar_coluna(ws, HEADER_COFINS_TOTAL)

    proximo_header = ws.cell(row=1, column=idx_t + 1).value
    if proximo_header == "PIS %":
        raise ValueError(
            "Esta planilha já contém as colunas de imposto (PIS %, COFINS %, PIS 2,1, "
            "COFINS 9,65, PIS, COFINS) logo após \"COFINS Total(R$)\" — nada foi alterado."
        )

    letra_p = get_column_letter(idx_p)
    letra_s = get_column_letter(idx_s)
    letra_t = get_column_letter(idx_t)
    formato_moeda = ws.cell(row=2, column=idx_s).number_format or "#,##0.00"
    header_font = copy(ws.cell(row=1, column=idx_t).font)

    insert_at = idx_t + 1
    ws.insert_cols(insert_at, amount=6)
    col_u, col_v, col_w, col_x, col_y, col_z = range(insert_at, insert_at + 6)
    letra_w = get_column_letter(col_w)
    letra_x = get_column_letter(col_x)

    definicoes = [
        (col_u, "PIS %", FILL_PERCENTUAL, FORMATO_PERCENTUAL, lambda r: f"={letra_s}{r}/{letra_p}{r}"),
        (col_v, "COFINS %", FILL_PERCENTUAL, FORMATO_PERCENTUAL, lambda r: f"={letra_t}{r}/{letra_p}{r}"),
        (col_w, "PIS 2,1", FILL_ALIQUOTA, formato_moeda, lambda r: f"={letra_p}{r}*0.021"),
        (col_x, "COFINS 9,65", FILL_ALIQUOTA, formato_moeda, lambda r: f"={letra_p}{r}*0.0965"),
        (col_y, "PIS", FILL_RESULTADO, formato_moeda, lambda r: f"={letra_w}{r}-{letra_s}{r}"),
        (col_z, "COFINS", FILL_RESULTADO, formato_moeda, lambda r: f"={letra_x}{r}-{letra_t}{r}"),
    ]

    total_pis = 0.0
    total_cofins = 0.0
    for col_idx, header, fill, number_format, formula_fn in definicoes:
        header_cell = ws.cell(row=1, column=col_idx, value=header)
        header_cell.font = header_font
        header_cell.fill = fill
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx, value=formula_fn(r))
            cell.number_format = number_format

    for r in range(2, ws.max_row + 1):
        p = ws.cell(row=r, column=idx_p).value or 0
        s = ws.cell(row=r, column=idx_s).value or 0
        t = ws.cell(row=r, column=idx_t).value or 0
        total_pis += (p * 0.021) - s
        total_cofins += (p * 0.0965) - t

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), {
        "linhas": ws.max_row - 1,
        "total_pis": round(total_pis, 2),
        "total_cofins": round(total_cofins, 2),
    }

"""
Geração do Excel de apuração do DIFAL de compras na Bahia.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .xml_parser import ItemDifal
from .calculator import ResultadoDifalItem

HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

COLUNAS = [
    ("Arquivo", 24), ("Chave NF-e", 26), ("Número NF", 10), ("Emissão", 12),
    ("CNPJ Emitente", 16), ("Emitente", 34), ("UF Origem", 8), ("UF Destino", 8),
    ("Regime", 16), ("CFOP", 8), ("Item", 6), ("Descrição", 32),
    ("Valor Operação", 14), ("Alíq. Interestadual", 12), ("Alíq. Interestadual Ref.", 12),
    ("ICMS Interestadual", 14), ("Base Reduzida", 14), ("Alíq. Interna BA", 12), ("Base Reajustada", 14),
    ("Diferença Alíquotas", 12), ("DIFAL Devido", 14), ("Fórmula", 10), ("Observações", 44),
]

COLS_MOEDA = {13, 16, 17, 19, 21}
COLS_PERCENTUAL = {14, 15, 18, 20}


def _observacoes(item: ItemDifal) -> str:
    obs = []
    if not item.icms_destacado:
        if item.percentual_credito_simples is not None:
            obs.append(
                f"Simples Nacional (CSOSN com permissão de crédito) — extraído {item.percentual_credito_simples:.2f}% "
                f"de ICMS embutido no preço (art. 23 da LC 123/2006); diferença calculada com a alíquota "
                f"interestadual constitucional de {item.aliquota_interestadual_referencia:.2%} "
                f"(Res. Senado 22/89{'/13/2012' if item.aliquota_interestadual_referencia == 0.04 else ''})."
            )
        else:
            obs.append(
                f"Simples Nacional sem percentual de crédito de ICMS informado na nota (CSOSN sem permissão de "
                f"crédito, ou campo não preenchido) — ICMS embutido assumido em 0% (posição conservadora); "
                f"diferença calculada com a alíquota interestadual constitucional de "
                f"{item.aliquota_interestadual_referencia:.2%} (Res. Senado 22/89"
                f"{'/13/2012' if item.aliquota_interestadual_referencia == 0.04 else ''}), independente do regime "
                f"do remetente."
            )
    if item.substituicao_tributaria:
        obs.append("ICMS-ST identificado no XML — fórmula de DIFAL-ST aplicada")
    if item.reducao_base_convenio_5291:
        obs.append(
            f"Redução de base do ICMS ao amparo do Convênio ICMS 52/91 (pRedBC "
            f"{item.percentual_reducao_bc:.2f}%) — DIFAL calculado com a alíquota interna "
            f"reduzida de 5,60% reconhecida pela Bahia (Art. 266 do RICMS-BA), em vez dos "
            f"20,5% padrão. Confirme a destinação industrial do bem, pois o fisco baiano "
            f"restringe este benefício a uso industrial do adquirente."
        )
    elif item.percentual_reducao_bc:
        obs.append(
            f"Redução de base do ICMS na origem (pRedBC {item.percentual_reducao_bc:.2f}%) não "
            f"identificada como Convênio 52/91 — a Bahia, em regra, não reconhece essa redução "
            f"para fins de DIFAL; calculado com a alíquota interna cheia de 20,5%. Verifique "
            f"manualmente se algum benefício específico se aplica."
        )
    if item.uf_destino != "BA":
        obs.append(f"ATENÇÃO: UF de destino da nota é {item.uf_destino}, não BA")
    if item.uf_origem == "BA":
        obs.append("ATENÇÃO: operação interna (origem também BA) — DIFAL pode não se aplicar")
    if item.ind_final != "1":
        obs.append("ATENÇÃO: indFinal ≠ 1 — confirmar se é operação de uso/consumo/ativo")
    return "; ".join(obs)


def gerar_excel(linhas: list[tuple[ItemDifal, ResultadoDifalItem]], avisos: list[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "DIFAL Bahia"

    for col, (nome, largura) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col, value=nome)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = largura
    ws.freeze_panes = "A2"

    row = 2
    total_difal = 0.0
    for item, res in linhas:
        valores = [
            item.arquivo, item.chave_nfe, item.numero_nf, (item.data_emissao or "")[:10],
            item.cnpj_emitente, item.nome_emitente, item.uf_origem, item.uf_destino,
            item.regime_emitente, item.cfop, item.n_item, item.descricao_produto,
            round(item.valor_operacao, 2), res.aliquota_interestadual, res.aliquota_interestadual_referencia,
            round(res.icms_interestadual, 2), round(res.base_reduzida, 2), res.aliquota_interna,
            round(res.base_reajustada, 2), res.diferenca_aliquotas, round(res.difal, 2),
            res.formula.upper(), _observacoes(item),
        ]
        for col, val in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            if col in COLS_MOEDA:
                cell.number_format = "#,##0.00"
            if col in COLS_PERCENTUAL:
                cell.number_format = "0.00%"
        total_difal += res.difal
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value=f"TOTAL — {len(linhas)} item(ns)").font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=20)
    tot_cell = ws.cell(row=total_row, column=21, value=round(total_difal, 2))
    tot_cell.font = Font(bold=True)
    tot_cell.number_format = "#,##0.00"

    if avisos:
        ws2 = wb.create_sheet("Avisos")
        ws2.cell(row=1, column=1, value="Avisos do processamento").font = Font(bold=True, size=12)
        for i, aviso in enumerate(avisos, start=3):
            ws2.cell(row=i, column=1, value=aviso)
        ws2.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

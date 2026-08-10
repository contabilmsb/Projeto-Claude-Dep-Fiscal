"""
Geração do Excel consolidado de Retenções CSRF.
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

LARGURAS = {
    "Código do Fornecedor": 16,
    "Nome/Razão Social do Fornecedor": 40,
    "CNPJ do Fornecedor": 18,
    "Data do Arquivo PCC": 14,
    "Número da Nota Fiscal": 16,
    "Código do Imposto Retido na Fonte": 18,
    "Origem do Valor": 14,
    "Valor do Imposto Retido na Fonte": 18,
    "Comprovante": 16,
    "Comprovante de Pagamento": 22,
}
COLS_MOEDA = {"Origem do Valor", "Valor do Imposto Retido na Fonte"}
COLS_DATA = {"Data do Arquivo PCC"}

LARGURAS_ACUMULADO = {
    "Código do Fornecedor": 16,
    "Nome/Razão Social do Fornecedor": 40,
    "CNPJ do Fornecedor": 18,
    "Data do Arquivo PCC (mais recente)": 22,
    "COFINS Retido": 14,
    "CSLL Retido": 14,
    "PIS Retido": 14,
    "Total Retido": 14,
}
COLS_MOEDA_ACUMULADO = {"COFINS Retido", "CSLL Retido", "PIS Retido", "Total Retido"}
COLS_DATA_ACUMULADO = {"Data do Arquivo PCC (mais recente)"}


def _valor_celula(v):
    if v is None or (isinstance(v, float) and pd.isna(v)) or v is pd.NaT:
        return None
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    return v


def _escrever_planilha(ws, df: pd.DataFrame, larguras: dict, cols_moeda: set, cols_data: set,
                        coluna_total: str | None = None) -> None:
    colunas = list(df.columns)
    for col_idx, nome in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(nome, 16)
    ws.freeze_panes = "A2"

    row_idx = 2
    for row in df.itertuples(index=False):
        for col_idx, (nome, valor) in enumerate(zip(colunas, row), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_valor_celula(valor))
            if nome in cols_moeda:
                cell.number_format = "#,##0.00"
            if nome in cols_data:
                cell.number_format = "dd/mm/yyyy"
        row_idx += 1

    if coluna_total:
        total_row = row_idx
        idx_valor = colunas.index(coluna_total) + 1
        ws.cell(row=total_row, column=1, value=f"TOTAL — {len(df)} linha(s)").font = Font(bold=True)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=idx_valor - 1)
        total = float(pd.to_numeric(df[coluna_total], errors="coerce").fillna(0).sum())
        tot_cell = ws.cell(row=total_row, column=idx_valor, value=round(total, 2))
        tot_cell.font = Font(bold=True)
        tot_cell.number_format = "#,##0.00"


def gerar_excel(df: pd.DataFrame, df_acumulado: pd.DataFrame, avisos: list[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Retenções CSRF"
    _escrever_planilha(ws, df, LARGURAS, COLS_MOEDA, COLS_DATA, coluna_total="Valor do Imposto Retido na Fonte")

    ws_acum = wb.create_sheet("Acumulado por Fornecedor")
    _escrever_planilha(
        ws_acum, df_acumulado, LARGURAS_ACUMULADO, COLS_MOEDA_ACUMULADO, COLS_DATA_ACUMULADO,
        coluna_total="Total Retido",
    )

    if avisos:
        ws2 = wb.create_sheet("Avisos")
        ws2.cell(row=1, column=1, value="Avisos do processamento").font = Font(bold=True, size=12)
        for i, aviso in enumerate(avisos, start=3):
            ws2.cell(row=i, column=1, value=aviso)
        ws2.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

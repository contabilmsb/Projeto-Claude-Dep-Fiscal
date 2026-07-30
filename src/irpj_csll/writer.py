"""
Geração da planilha de saída do módulo IRPJ/CSLL.

O template "APURAÇÃO IRPJ CSLL.xlsx" é protegido por senha (a contabilidade
mantém um único arquivo, com um bloco por trimestre, sempre no mesmo
layout). Este módulo:
  1. Decripta o template em memória via msoffcrypto-tool (a senha nunca é
     gravada no código — vem da variável de ambiente
     IRPJ_CSLL_TEMPLATE_PASSWORD).
  2. Localiza o bloco do trimestre pedido (procurando o texto
     "Nº Trimestre AAAA") nas abas IRPJ e CSLL.
  3. Preenche os valores calculados nas colunas dos 3 meses e nas linhas de
     totais/deduções/parcelas, por busca de rótulo — nunca por posição fixa
     de célula, para não quebrar se a contabilidade inserir linhas.

Se o bloco do trimestre ainda não existir no template (ex.: um trimestre
novo que a contabilidade ainda não criou), uma exceção clara é levantada —
a apuração em si (dashboard) não depende deste arquivo para funcionar.
"""

import io
import os
import re
import datetime
from pathlib import Path

import msoffcrypto
from openpyxl import load_workbook
from openpyxl.styles import Font

from .calculator import ComponenteMes, ResultadoTrimestre

FONT_NAME = "Arial"
COLOR_INPUT = "0000FF"
COLOR_FORMULA = "000000"
TEMPLATE_PASSWORD_ENV = "IRPJ_CSLL_TEMPLATE_PASSWORD"


def _decrypt_to_stream(path: Path) -> io.BytesIO:
    senha = os.getenv(TEMPLATE_PASSWORD_ENV, "")
    with open(path, "rb") as f:
        raw = f.read()
    buf = io.BytesIO(raw)
    try:
        office_file = msoffcrypto.OfficeFile(buf)
        office_file.load_key(password=senha)
        decrypted = io.BytesIO()
        office_file.decrypt(decrypted)
        decrypted.seek(0)
        return decrypted
    except Exception:
        # Não está protegido (ou senha não configurada) — usa o conteúdo original
        return io.BytesIO(raw)


def _find_trimestre_header(sheet, trimestre: int, ano: int) -> int | None:
    alvo = f"{trimestre}º trimestre {ano}".lower()
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and alvo in str(cell.value).lower().strip():
                return cell.row
    return None


_BLOCO_HEADER_RE = re.compile(r"\d+º\s*trimestre\s*\d{4}", re.IGNORECASE)


def _next_block_start(sheet, header_row: int) -> int:
    """
    Linha onde termina o bloco atual (início do próximo bloco "Nº Trimestre AAAA"
    ou fim da aba). Usa o padrão completo (não apenas a palavra "trimestre") para
    não confundir com rótulos como "VALOR DO IRPJ DEVIDO NO TRIMESTRE".
    """
    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            if cell.value and _BLOCO_HEADER_RE.search(str(cell.value)):
                return cell.row
    return sheet.max_row + 1


def _month_columns(sheet, header_row: int) -> list[int]:
    """Colunas dos 3 meses — linha logo abaixo do cabeçalho do bloco, valores de data serial."""
    cols = []
    data_row = header_row + 1
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=data_row, column=col).value
        if isinstance(val, datetime.datetime) or (isinstance(val, (int, float)) and val > 40000):
            cols.append(col)
    return cols[:3]


def _find_row_by_label(sheet, label_partial: str, row_start: int, row_end: int,
                        label_cols: range = range(1, 8)) -> int | None:
    label_lower = label_partial.lower().strip()
    for r in range(row_start, row_end):
        for c in label_cols:
            val = sheet.cell(row=r, column=c).value
            if val and label_lower in str(val).lower():
                return r
    return None


def _write_row(sheet, row: int, month_cols: list[int], total_col: int, valores: list[float]):
    for col, valor in zip(month_cols, valores):
        cell = sheet.cell(row=row, column=col)
        cell.value = round(valor, 2)
        cell.font = Font(name=FONT_NAME, color=COLOR_INPUT, size=10)
        cell.number_format = "#,##0.00"
    if total_col:
        tot = sheet.cell(row=row, column=total_col)
        tot.value = round(sum(valores), 2)
        tot.font = Font(name=FONT_NAME, color=COLOR_FORMULA, size=10, bold=True)
        tot.number_format = "#,##0.00"


def _preencher_bloco(sheet, trimestre: int, ano: int, meses: list[ComponenteMes],
                      resultado_tributo, deducao_labels: dict, base_calc_label: str,
                      valor_a_recolher_label: str, resultado_col: int):
    """
    `resultado_col` é a coluna fixa de "BASE CALC" do bloco (N=14 na aba IRPJ,
    M=13 na aba CSLL) — não deve ser confundida com a coluna "TOTAL" (soma dos
    3 meses), que fica logo após a última coluna de mês.
    """
    header_row = _find_trimestre_header(sheet, trimestre, ano)
    if header_row is None:
        raise ValueError(
            f"Bloco '{trimestre}º Trimestre {ano}' não encontrado no template "
            f"(aba '{sheet.title}'). Adicione o bloco desse trimestre na planilha "
            "antes de exportar, ou verifique se a competência está correta."
        )
    row_end = _next_block_start(sheet, header_row)
    month_cols = _month_columns(sheet, header_row)
    total_col = month_cols[-1] + 1 if month_cols else None

    def escreve(label: str, valores: list[float]):
        row = _find_row_by_label(sheet, label, header_row, row_end)
        if row:
            _write_row(sheet, row, month_cols, total_col, valores)

    escreve("vendas/exporta", [m.revenda_base for m in meses])
    escreve("rendto de aplic financeira", [m.aplicacao_financeira for m in meses])
    escreve("var cambial ativa", [m.variacao_cambial for m in meses])
    escreve("juros ativos", [m.juros_recebidos for m in meses])

    for label, valores in deducao_labels.items():
        escreve(label, valores)

    row_base = _find_row_by_label(sheet, base_calc_label, header_row, row_end)
    if row_base:
        cell = sheet.cell(row=row_base, column=resultado_col)
        cell.value = round(resultado_tributo.base_calculo, 2)
        cell.font = Font(name=FONT_NAME, bold=True, size=10)
        cell.number_format = "#,##0.00"

    row_pagar = _find_row_by_label(sheet, valor_a_recolher_label, header_row, row_end)
    if row_pagar:
        cell = sheet.cell(row=row_pagar, column=resultado_col)
        cell.value = round(resultado_tributo.valor_a_pagar, 2)
        cell.font = Font(name=FONT_NAME, bold=True, size=11)
        cell.number_format = "#,##0.00"


def atualizar_template(template_path: Path, output_path: Path,
                        resultado: ResultadoTrimestre, meses: list[ComponenteMes]) -> Path:
    stream = _decrypt_to_stream(template_path)
    wb = load_workbook(stream)

    if "IRPJ" in wb.sheetnames:
        _preencher_bloco(
            wb["IRPJ"], resultado.trimestre, resultado.ano, meses, resultado.irpj,
            deducao_labels={
                "ir retido fonte por org": [m.irrf_cliente for m in meses],
                "ir retido fonte s/rendt aplica": [m.irrf_aplicacao for m in meses],
            },
            base_calc_label="base de calculo para irpj",
            valor_a_recolher_label="valor  irpj a recolher",
            resultado_col=14,  # coluna N — "BASE CALC" na aba IRPJ
        )

    if "CSLL" in wb.sheetnames:
        _preencher_bloco(
            wb["CSLL"], resultado.trimestre, resultado.ano, meses, resultado.csll,
            deducao_labels={
                "csl retido fonte por org": [m.csll_retida for m in meses],
            },
            base_calc_label="base de calculo para",
            valor_a_recolher_label="valor csll a recolher",
            resultado_col=13,  # coluna M — "BASE CALC" na aba CSLL
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path

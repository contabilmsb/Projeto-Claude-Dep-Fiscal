"""
Geração do Excel da tabela "Recebimento - Notas Fiscais" exatamente como
exibida na tela (mesmas linhas visíveis, na mesma ordem — respeita o filtro
de busca e a ordenação de coluna aplicados pelo usuário).
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

COLUNAS = [
    ("competencia", "Competência", 14, None),
    ("nf", "NF", 14, None),
    ("cliente", "Cliente", 40, None),
    ("recebido", "Recebido", 16, "#,##0.00"),
    ("cofins_retido", "COFINS Ret.", 14, "#,##0.00"),
    ("pis_retido", "PIS Ret.", 14, "#,##0.00"),
    ("csll_retido", "CSLL Ret.", 14, "#,##0.00"),
    ("irrf", "IRRF", 14, "#,##0.00"),
    ("juros", "Juros", 14, "#,##0.00"),
    ("base_liquida", "Base Líquida", 16, "#,##0.00"),
]
COLUNAS_SOMA = {"recebido", "cofins_retido", "pis_retido", "csll_retido", "irrf", "juros", "base_liquida"}


def gerar_excel_consolidacao_nf(linhas: list[dict], competencia_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Recebimento - Notas Fiscais"

    for col_idx, (_, titulo, largura, _) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.freeze_panes = "A2"

    row_idx = 2
    for linha in linhas:
        for col_idx, (chave, _, _, formato) in enumerate(COLUNAS, start=1):
            valor = linha.get(chave)
            cell = ws.cell(row=row_idx, column=col_idx, value=valor if valor not in (None, "") else None)
            if formato:
                cell.number_format = formato
        row_idx += 1

    total_row = row_idx
    ws.cell(row=total_row, column=1, value=f"TOTAL — {len(linhas)} NFs — {competencia_label}").font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    for col_idx, (chave, _, _, formato) in enumerate(COLUNAS, start=1):
        if chave not in COLUNAS_SOMA:
            continue
        total = sum(float(linha.get(chave) or 0) for linha in linhas)
        tot_cell = ws.cell(row=total_row, column=col_idx, value=round(total, 2))
        tot_cell.font = Font(bold=True)
        tot_cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
